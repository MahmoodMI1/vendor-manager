import openpyxl
from datetime import date, datetime

# --- COLUMN HEADERS (edit these to match the customer's Excel) ---
SCHEDULE_NAME_HEADER = "Vendor Name"
SCHEDULE_DATE_HEADER = "Visit Date"
DIRECTORY_NAME_HEADER = "Vendor Name"
DIRECTORY_EMAIL_HEADER = "Email"
# -----------------------------------------------------------------


def _normalize_date(value) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        value = value.strip()
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y"):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
        return None
    if isinstance(value, (int, float)):
        try:
            return datetime.fromordinal(datetime(1899, 12, 30).toordinal() + int(value)).date()
        except (ValueError, OverflowError):
            return None
    return None


def _find_columns(sheet, headers: list[str]) -> dict[str, int]:
    first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    mapping = {}
    for target in headers:
        for i, cell in enumerate(first_row):
            if cell and str(cell).strip().lower() == target.lower():
                mapping[target] = i
                break
        if target not in mapping:
            raise ValueError(f'Column "{target}" not found. Found: {[str(c) for c in first_row if c]}')
    return mapping


def read_schedule(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    cols = _find_columns(ws, [SCHEDULE_NAME_HEADER, SCHEDULE_DATE_HEADER])
    name_idx = cols[SCHEDULE_NAME_HEADER]
    date_idx = cols[SCHEDULE_DATE_HEADER]

    rows = []
    first = True
    for row in ws.iter_rows(values_only=True):
        if first:
            first = False
            continue

        name = row[name_idx] if name_idx < len(row) else None
        raw_date = row[date_idx] if date_idx < len(row) else None

        if not name or not str(name).strip():
            continue

        parsed_date = _normalize_date(raw_date)
        if parsed_date is None:
            continue

        rows.append({
            "vendor_name": str(name).strip(),
            "visit_date": parsed_date,
        })

    wb.close()
    return rows


def read_directory(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    cols = _find_columns(ws, [DIRECTORY_NAME_HEADER, DIRECTORY_EMAIL_HEADER])
    name_idx = cols[DIRECTORY_NAME_HEADER]
    email_idx = cols[DIRECTORY_EMAIL_HEADER]

    rows = []
    first = True
    for row in ws.iter_rows(values_only=True):
        if first:
            first = False
            continue

        name = row[name_idx] if name_idx < len(row) else None
        email = row[email_idx] if email_idx < len(row) else None

        if not name or not str(name).strip():
            continue
        if not email or not str(email).strip():
            continue

        rows.append({
            "vendor_name": str(name).strip(),
            "email": str(email).strip(),
        })

    wb.close()
    return rows


def get_visits_for_date(schedule: list[dict], target: date) -> list[dict]:
    return [row for row in schedule if row["visit_date"] == target]


def lookup_vendor_email(vendor_name: str, directory: list[dict]) -> str | None:
    target = vendor_name.lower()
    for entry in directory:
        if entry["vendor_name"].lower() == target:
            return entry["email"]
    return None